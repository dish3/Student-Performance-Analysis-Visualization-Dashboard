# ============================================================
# app.py  –  Main Flask Application
# ============================================================
# This is the entry point of the Student Performance Analysis
# and Visualization Dashboard (SPAVD).
# It defines all URL routes, handles HTTP requests/responses,
# talks to the MySQL database, and renders HTML templates.
#
# Route groups:
#   /              – redirect to dashboard or login
#   /login         – authentication
#   /logout        – session clear
#   /admin/...     – admin-only pages
#   /faculty/...   – faculty + admin pages
# ============================================================

# ── Imports ──────────────────────────────────────────────────
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, flash, send_file, jsonify
)
# render_template  – renders an HTML template file from /templates
# request          – access form data, files, query params
# redirect         – send the browser to a different URL
# url_for          – build a URL from a function name (avoids hardcoding)
# session          – server-side cookie dict (stores login state)
# flash            – one-time messages shown on the next page
# send_file        – stream a file download to the browser

from werkzeug.security import check_password_hash, generate_password_hash
# check_password_hash  – safely compare a plain password to a stored hash
# generate_password_hash – hash a plain password before storing it

import os
import mysql.connector          # MySQL driver for Python

from openpyxl import Workbook                          # used for the raw-data Excel export
from openpyxl.styles import Font, PatternFill, Alignment  # cell styling for that export

from datetime import datetime   # used to timestamp export filenames
import io                       # BytesIO – in-memory file buffer (no temp files on disk)
import excel_manager            # our custom module for analytics export + import
import analytics               # SQL query functions for advanced analytics
import analytics_export        # pandas + xlsxwriter multi-sheet export


# ── App setup ─────────────────────────────────────────────────
app = Flask(__name__)
# secret_key is required for session encryption.
# Change this to a long random string in production!
app.secret_key = os.environ.get('APP_SECRET_KEY') or \
    os.environ.get('SECRET_KEY') or \
    'your-secret-key-change-this-in-production'


def get_env(*names, default=None):
    """Return the first non-empty environment variable from names."""
    for name in names:
        value = os.environ.get(name)
        if value:
            return value
    return default


DB_CONFIG = {
    'host': get_env('DB_HOST', 'MYSQL_HOST', default='localhost'),
    'port': int(get_env('DB_PORT', 'MYSQL_PORT', default='3306')),
    'user': get_env('DB_USER', 'MYSQL_USER', default='root'),
    'password': get_env('DB_PASSWORD', 'MYSQL_PASSWORD', default='root'),
    'database': get_env(
        'DB_NAME', 'MYSQL_DATABASE', 'MYSQL_DB',
        default='student_assessment_system'
    )
}


# ============================================================
#  HELPER FUNCTIONS
# ============================================================

def get_db_connection():
    """
    Open and return a new MySQL connection using environment settings.
    Returns None if the connection fails (caller must handle this).
    """
    try:
        conn = mysql.connector.connect(**DB_CONFIG)  # unpack dict as keyword args
        return conn
    except mysql.connector.Error as e:
        safe_db_config = {
            'host': DB_CONFIG.get('host'),
            'port': DB_CONFIG.get('port'),
            'user': DB_CONFIG.get('user'),
            'database': DB_CONFIG.get('database')
        }
        print(f"Database connection error: {e}")  # log to console for debugging
        print(f"Database settings in use: {safe_db_config}")
        return None


def is_logged_in():
    """
    Return True if the current browser session has a valid login.
    Checks that 'logged_in' key exists AND equals True in the session dict.
    """
    return 'logged_in' in session and session['logged_in'] == True


def is_admin():
    """
    Return True if the logged-in user has the 'admin' role.
    Admins can access all routes; faculty can only access marks routes.
    """
    return is_logged_in() and session.get('role') == 'admin'


# ============================================================
#  AUTHENTICATION ROUTES
# ============================================================

@app.route('/')
def index():
    """
    Root URL – redirect to the correct dashboard based on role,
    or to the login page if not logged in.
    """
    if is_logged_in():
        if is_admin():
            return redirect(url_for('admin_dashboard'))   # admin → admin dashboard
        else:
            return redirect(url_for('faculty_dashboard')) # faculty → faculty dashboard
    return redirect(url_for('login'))  # not logged in → login page


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    GET  – show the login form.
    POST – validate credentials against the faculty table.
           Both admins and faculty use the same login form;
           the 'role' column in the DB determines what they see.
    """
    if is_logged_in():
        return redirect(url_for('index'))  # already logged in, skip login page

    if request.method == 'POST':
        username = request.form.get('username')  # from the HTML <input name="username">
        password = request.form.get('password')  # from the HTML <input name="password">

        if not username or not password:
            flash('Please enter both username and password', 'danger')
            return render_template('login.html')

        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)  # dictionary=True → rows as dicts
            # Look up the user by username (username is UNIQUE in the DB)
            cursor.execute("SELECT * FROM faculty WHERE username = %s", (username,))
            user = cursor.fetchone()  # returns dict or None
            cursor.close()
            conn.close()

            # check_password_hash compares the plain password to the stored bcrypt hash
            if user and check_password_hash(user['password'], password):
                # Store login info in the session (encrypted cookie)
                session['logged_in']  = True
                session['username']   = username
                session['faculty_id'] = user['faculty_id']
                session['role']       = user['role']  # 'admin' or 'faculty'
                flash('Login successful!', 'success')
                return redirect(url_for('index'))
            else:
                flash('Invalid username or password', 'danger')
        else:
            flash('Database connection error', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    """
    Clear the session (removes all stored login data) and
    redirect to the login page.
    """
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))


# ============================================================
#  DASHBOARD ROUTES
# ============================================================

@app.route('/admin/dashboard')
def admin_dashboard():
    """
    Admin dashboard – shows 5 summary statistics:
    total students, tests, attempts, attendance records, faculty count.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('login'))

    cursor = conn.cursor()

    # Count total students in the students table
    cursor.execute("SELECT COUNT(*) FROM students")
    total_students = cursor.fetchone()[0]

    # Count total tests created
    cursor.execute("SELECT COUNT(*) FROM tests")
    total_tests = cursor.fetchone()[0]

    # Count unique student+test+attempt combinations (= number of test sittings)
    cursor.execute("SELECT COUNT(DISTINCT moodle_id, test_id, attempt_no) FROM student_test_marks")
    total_attempts = cursor.fetchone()[0]

    # Count total daily attendance entries across all students
    cursor.execute("SELECT COUNT(*) FROM attendance_daily")
    total_attendance = cursor.fetchone()[0]

    # Count faculty accounts (excludes admins)
    cursor.execute("SELECT COUNT(*) FROM faculty WHERE role = 'faculty'")
    total_faculty = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    # Pass all counts to the template as template variables
    return render_template('admin_dashboard.html',
                           total_students=total_students,
                           total_tests=total_tests,
                           total_attempts=total_attempts,
                           total_attendance=total_attendance,
                           total_faculty=total_faculty)


@app.route('/faculty/dashboard')
def faculty_dashboard():
    """
    Faculty dashboard – shows 3 summary statistics.
    Admins are redirected to their own dashboard.
    """
    if not is_logged_in():
        flash('Please login first', 'danger')
        return redirect(url_for('login'))

    if is_admin():
        return redirect(url_for('admin_dashboard'))  # admins don't use this page

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('login'))

    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM students")
    total_students = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM tests")
    total_tests = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(DISTINCT moodle_id, test_id, attempt_no) FROM student_test_marks")
    total_attempts = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    return render_template('faculty_dashboard.html',
                           total_students=total_students,
                           total_tests=total_tests,
                           total_attempts=total_attempts)


# ============================================================
#  STUDENT MANAGEMENT ROUTES  (admin only)
# ============================================================

@app.route('/admin/add_student', methods=['GET', 'POST'])
def add_student():
    """
    GET  – show the Add Student form.
    POST – validate input and INSERT a new row into `students`.
           Also inserts a default 0% row in `attendance_percentage`.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        # Read all 5 fields from the submitted form
        moodle_id    = request.form.get('moodle_id')
        roll_no      = request.form.get('roll_no')
        name         = request.form.get('name')
        moodle_email = request.form.get('moodle_email')
        phone        = request.form.get('phone')

        # all() returns False if any value is falsy (None or empty string)
        if not all([moodle_id, roll_no, name, moodle_email, phone]):
            flash('All fields are required', 'danger')
            return render_template('add_student.html')

        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                # Insert the new student; moodle_id is the PRIMARY KEY
                cursor.execute(
                    "INSERT INTO students (moodle_id, roll_no, name, moodle_email, phone) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (moodle_id, roll_no, name, moodle_email, phone)
                )
                conn.commit()

                # Create a matching attendance_percentage row defaulting to 0%
                cursor.execute(
                    "INSERT INTO attendance_percentage (moodle_id, attendance_percent) VALUES (%s, 0)",
                    (moodle_id,)
                )
                conn.commit()

                flash('Student added successfully!', 'success')
                cursor.close()
                conn.close()
                return redirect(url_for('view_students'))

            except mysql.connector.IntegrityError:
                # Triggered when moodle_id, roll_no, or email already exists (UNIQUE constraint)
                flash('Error: Moodle ID, Roll No, or Email already exists', 'danger')
                cursor.close()
                conn.close()
            except Exception as e:
                flash(f'Error adding student: {str(e)}', 'danger')
                cursor.close()
                conn.close()
        else:
            flash('Database connection error', 'danger')

    return render_template('add_student.html')


@app.route('/admin/edit_student/<moodle_id>', methods=['GET', 'POST'])
def edit_student(moodle_id):
    """
    GET  – load the student's current data and show the edit form.
    POST – validate and UPDATE the student's row (moodle_id cannot change).
    <moodle_id> is captured from the URL, e.g. /admin/edit_student/M001
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('view_students'))

    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        roll_no      = request.form.get('roll_no')
        name         = request.form.get('name')
        moodle_email = request.form.get('moodle_email')
        phone        = request.form.get('phone')

        if not all([roll_no, name, moodle_email, phone]):
            flash('All fields are required', 'danger')
            # Re-fetch student to re-populate the form
            cursor.execute("SELECT * FROM students WHERE moodle_id = %s", (moodle_id,))
            student = cursor.fetchone()
            cursor.close()
            conn.close()
            return render_template('edit_student.html', student=student)

        try:
            # UPDATE the 4 editable fields; moodle_id stays the same
            cursor.execute(
                "UPDATE students SET roll_no=%s, name=%s, moodle_email=%s, phone=%s "
                "WHERE moodle_id=%s",
                (roll_no, name, moodle_email, phone, moodle_id)
            )
            conn.commit()
            flash('Student updated successfully!', 'success')
            cursor.close()
            conn.close()
            return redirect(url_for('view_students'))

        except mysql.connector.IntegrityError:
            flash('Error: Roll No or Email already exists', 'danger')
            cursor.execute("SELECT * FROM students WHERE moodle_id = %s", (moodle_id,))
            student = cursor.fetchone()
            cursor.close()
            conn.close()
            return render_template('edit_student.html', student=student)

        except Exception as e:
            flash(f'Error updating student: {str(e)}', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('view_students'))

    # GET – fetch the student's current data to pre-fill the form
    cursor.execute("SELECT * FROM students WHERE moodle_id = %s", (moodle_id,))
    student = cursor.fetchone()
    cursor.close()
    conn.close()

    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('view_students'))

    return render_template('edit_student.html', student=student)


@app.route('/admin/view_students')
def view_students():
    """Fetch all students ordered by moodle_id and display them in a table."""
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM students ORDER BY moodle_id")
    students = cursor.fetchall()  # list of dicts
    cursor.close()
    conn.close()

    return render_template('view_students.html', students=students)


@app.route('/admin/delete_student/<moodle_id>')
def delete_student(moodle_id):
    """
    DELETE a student by moodle_id.
    ON DELETE CASCADE in the DB schema automatically removes all
    related marks, attendance, and attendance_percentage rows.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM students WHERE moodle_id = %s", (moodle_id,))
            conn.commit()
            flash('Student deleted successfully!', 'success')
        except Exception as e:
            flash(f'Error deleting student: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    else:
        flash('Database connection error', 'danger')

    return redirect(url_for('view_students'))


# ============================================================
#  TEST MANAGEMENT ROUTES  (admin only)
# ============================================================

@app.route('/admin/create_test', methods=['GET', 'POST'])
def create_test():
    """
    GET  – show the Create Test form.
    POST – INSERT a new test name into `tests`, then redirect to
           add_test_questions so the admin can define questions.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        test_name = request.form.get('test_name')

        if not test_name:
            flash('Test name is required', 'danger')
            return render_template('create_test.html')

        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                cursor.execute("INSERT INTO tests (test_name) VALUES (%s)", (test_name,))
                conn.commit()
                test_id = cursor.lastrowid  # get the auto-generated test_id
                flash('Test created successfully! Now add questions.', 'success')
                cursor.close()
                conn.close()
                # Redirect to the question-setup page for this new test
                return redirect(url_for('add_test_questions', test_id=test_id))
            except Exception as e:
                flash(f'Error creating test: {str(e)}', 'danger')
                cursor.close()
                conn.close()
        else:
            flash('Database connection error', 'danger')

    return render_template('create_test.html')


@app.route('/admin/add_test_questions/<int:test_id>', methods=['GET', 'POST'])
def add_test_questions(test_id):
    """
    Two-step form for adding questions to a test:
      Step 1 – admin enters the number of questions.
      Step 2 – admin enters max marks for each question.

    The 'step' variable controls which part of the template is shown.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    # Fetch the test to display its name in the form
    cursor.execute("SELECT * FROM tests WHERE test_id = %s", (test_id,))
    test = cursor.fetchone()

    if not test:
        flash('Test not found', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('view_tests'))

    if request.method == 'POST':
        num_questions = request.form.get('num_questions')

        if not num_questions:
            flash('Number of questions is required', 'danger')
            cursor.close()
            conn.close()
            return render_template('add_test_questions.html', test=test, step=1)

        try:
            num_questions = int(num_questions)
            if num_questions <= 0:
                raise ValueError()
        except ValueError:
            flash('Please enter a valid number of questions', 'danger')
            cursor.close()
            conn.close()
            return render_template('add_test_questions.html', test=test, step=1)

        # If 'max_marks_1' is in the form, the admin submitted step 2 (marks per question)
        if 'max_marks_1' in request.form:
            try:
                # Loop through each question and insert into test_questions
                for i in range(1, num_questions + 1):
                    max_marks = request.form.get(f'max_marks_{i}')
                    if not max_marks or int(max_marks) <= 0:
                        raise ValueError(f'Invalid max marks for question {i}')

                    cursor.execute(
                        "INSERT INTO test_questions (test_id, question_number, max_marks) "
                        "VALUES (%s, %s, %s)",
                        (test_id, i, int(max_marks))
                    )

                conn.commit()
                flash(f'Successfully added {num_questions} questions!', 'success')
                cursor.close()
                conn.close()
                return redirect(url_for('view_tests'))

            except Exception as e:
                conn.rollback()  # undo partial inserts if any question failed
                flash(f'Error adding questions: {str(e)}', 'danger')
                cursor.close()
                conn.close()
                return render_template('add_test_questions.html', test=test, step=1)
        else:
            # Step 1 submitted – show step 2 (max marks input for each question)
            cursor.close()
            conn.close()
            return render_template('add_test_questions.html', test=test,
                                   step=2, num_questions=num_questions)

    # GET – show step 1 (enter number of questions)
    cursor.close()
    conn.close()
    return render_template('add_test_questions.html', test=test, step=1)


@app.route('/admin/view_tests')
def view_tests():
    """
    List all tests with a count of how many questions each has.
    Uses LEFT JOIN so tests with 0 questions still appear.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT t.*, COUNT(tq.question_id) AS num_questions
        FROM tests t
        LEFT JOIN test_questions tq ON t.test_id = tq.test_id
        GROUP BY t.test_id
        ORDER BY t.test_id DESC
    """)
    tests = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('view_tests.html', tests=tests)


@app.route('/admin/view_test_details/<int:test_id>')
def view_test_details(test_id):
    """Show a test's name and all its questions with max marks."""
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tests WHERE test_id = %s", (test_id,))
    test = cursor.fetchone()

    if not test:
        flash('Test not found', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('view_tests'))

    # Fetch all questions for this test, ordered by question number
    cursor.execute(
        "SELECT * FROM test_questions WHERE test_id = %s ORDER BY question_number",
        (test_id,)
    )
    questions = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('view_test_details.html', test=test, questions=questions)


@app.route('/admin/delete_test/<int:test_id>')
def delete_test(test_id):
    """
    DELETE a test. ON DELETE CASCADE removes all related
    test_questions and student_test_marks automatically.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM tests WHERE test_id = %s", (test_id,))
            conn.commit()
            flash('Test deleted successfully!', 'success')
        except Exception as e:
            flash(f'Error deleting test: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    else:
        flash('Database connection error', 'danger')

    return redirect(url_for('view_tests'))


# ============================================================
#  MARKS ENTRY ROUTES  (faculty + admin)
# ============================================================

@app.route('/faculty/enter_marks', methods=['GET', 'POST'])
def enter_marks():
    """
    Two-step marks entry:
      Step 1 – select a test and a student (and optionally an attempt number).
      Step 2 – enter obtained marks for each question.

    Existing marks for the selected attempt are pre-loaded so the
    user can edit them. New attempts are created by choosing a new
    attempt number.
    """
    if not is_logged_in():
        flash('Please login first', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        # Redirect to the correct dashboard based on role
        return redirect(url_for('faculty_dashboard' if not is_admin() else 'admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        test_id    = request.form.get('test_id')
        moodle_id  = request.form.get('moodle_id')
        attempt_no = request.form.get('attempt_no')

        if not test_id or not moodle_id:
            flash('Please select both test and student', 'danger')
            cursor.execute("SELECT * FROM tests")
            tests = cursor.fetchall()
            cursor.execute("SELECT * FROM students ORDER BY moodle_id")
            students = cursor.fetchall()
            cursor.close()
            conn.close()
            return render_template('enter_marks.html', tests=tests, students=students, step=1)

        # 'submit_marks' hidden field is present only when step 2 form is submitted
        if 'submit_marks' in request.form:
            try:
                attempt_no = int(attempt_no)

                # Get all questions for this test to iterate over them
                cursor.execute(
                    "SELECT * FROM test_questions WHERE test_id = %s ORDER BY question_number",
                    (test_id,)
                )
                questions = cursor.fetchall()

                for question in questions:
                    # Read the mark for this question from the form
                    obtained_marks = request.form.get(f'obtained_marks_{question["question_id"]}')
                    obtained_marks = 0 if obtained_marks is None else int(obtained_marks)

                    # Check if a mark already exists for this student+test+attempt+question
                    cursor.execute("""
                        SELECT id FROM student_test_marks
                        WHERE moodle_id=%s AND test_id=%s AND attempt_no=%s AND question_id=%s
                    """, (moodle_id, test_id, attempt_no, question['question_id']))
                    existing = cursor.fetchone()

                    if existing:
                        # UPDATE the existing mark
                        cursor.execute("""
                            UPDATE student_test_marks SET obtained_marks=%s
                            WHERE moodle_id=%s AND test_id=%s AND attempt_no=%s AND question_id=%s
                        """, (obtained_marks, moodle_id, test_id, attempt_no, question['question_id']))
                    else:
                        # INSERT a new mark row
                        cursor.execute("""
                            INSERT INTO student_test_marks
                            (moodle_id, test_id, attempt_no, question_id, obtained_marks)
                            VALUES (%s, %s, %s, %s, %s)
                        """, (moodle_id, test_id, attempt_no, question['question_id'], obtained_marks))

                conn.commit()
                flash('Marks saved successfully!', 'success')
                cursor.close()
                conn.close()
                return redirect(url_for('view_marks'))

            except Exception as e:
                conn.rollback()  # undo all mark inserts/updates if any failed
                flash(f'Error saving marks: {str(e)}', 'danger')
                cursor.close()
                conn.close()
                return redirect(url_for('enter_marks'))

        else:
            # Step 1 submitted – load the marks entry form (step 2)
            attempt_no = 1 if not attempt_no else int(attempt_no)

            cursor.execute("SELECT * FROM tests WHERE test_id = %s", (test_id,))
            test = cursor.fetchone()

            cursor.execute("SELECT * FROM students WHERE moodle_id = %s", (moodle_id,))
            student = cursor.fetchone()

            cursor.execute(
                "SELECT * FROM test_questions WHERE test_id = %s ORDER BY question_number",
                (test_id,)
            )
            questions = cursor.fetchall()

            # Load any existing marks for this attempt so the form is pre-filled
            cursor.execute("""
                SELECT question_id, obtained_marks FROM student_test_marks
                WHERE moodle_id=%s AND test_id=%s AND attempt_no=%s
            """, (moodle_id, test_id, attempt_no))
            # Build a dict: {question_id: obtained_marks} for easy template lookup
            existing_marks = {row['question_id']: row['obtained_marks'] for row in cursor.fetchall()}

            # Get list of attempt numbers already recorded for this student+test
            cursor.execute("""
                SELECT DISTINCT attempt_no FROM student_test_marks
                WHERE moodle_id=%s AND test_id=%s ORDER BY attempt_no
            """, (moodle_id, test_id))
            attempts = [row['attempt_no'] for row in cursor.fetchall()]

            cursor.close()
            conn.close()

            return render_template('enter_marks.html',
                                   test=test, student=student, questions=questions,
                                   existing_marks=existing_marks, attempt_no=attempt_no,
                                   attempts=attempts, step=2)

    # GET – show step 1: test and student selection dropdowns
    cursor.execute("SELECT * FROM tests")
    tests = cursor.fetchall()
    cursor.execute("SELECT * FROM students ORDER BY moodle_id")
    students = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('enter_marks.html', tests=tests, students=students, step=1)


@app.route('/faculty/view_marks')
def view_marks():
    """
    Show a summary table of all marks records.
    Each row = one student + one test + one attempt.
    Totals and percentage are calculated in Python after fetching.
    """
    if not is_logged_in():
        flash('Please login first', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('faculty_dashboard' if not is_admin() else 'admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    # Get all unique student+test+attempt combinations
    cursor.execute("""
        SELECT DISTINCT s.moodle_id, s.roll_no, s.name,
                        t.test_id, t.test_name, stm.attempt_no
        FROM student_test_marks stm
        JOIN students s ON stm.moodle_id = s.moodle_id
        JOIN tests t    ON stm.test_id   = t.test_id
        ORDER BY t.test_id DESC, s.moodle_id, stm.attempt_no
    """)
    records = cursor.fetchall()

    # For each record, calculate total max marks, total obtained, and percentage
    for record in records:
        cursor.execute("""
            SELECT SUM(tq.max_marks)       AS total_max,
                   SUM(stm.obtained_marks) AS total_obtained
            FROM student_test_marks stm
            JOIN test_questions tq ON stm.question_id = tq.question_id
            WHERE stm.moodle_id=%s AND stm.test_id=%s AND stm.attempt_no=%s
        """, (record['moodle_id'], record['test_id'], record['attempt_no']))
        totals = cursor.fetchone()
        record['total_max']      = totals['total_max']      or 0
        record['total_obtained'] = totals['total_obtained'] or 0
        # Avoid division by zero
        record['percentage'] = (
            round(record['total_obtained'] / record['total_max'] * 100, 2)
            if record['total_max'] > 0 else 0
        )

    cursor.close()
    conn.close()

    return render_template('view_marks.html', records=records)


@app.route('/faculty/view_marks_detail/<moodle_id>/<int:test_id>/<int:attempt_no>')
def view_marks_detail(moodle_id, test_id, attempt_no):
    """
    Show question-by-question marks for one student, one test, one attempt.
    URL parameters: moodle_id, test_id, attempt_no
    """
    if not is_logged_in():
        flash('Please login first', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('faculty_dashboard' if not is_admin() else 'admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM students WHERE moodle_id = %s", (moodle_id,))
    student = cursor.fetchone()

    cursor.execute("SELECT * FROM tests WHERE test_id = %s", (test_id,))
    test = cursor.fetchone()

    # LEFT JOIN so questions with no mark entry still appear (with NULL obtained_marks)
    cursor.execute("""
        SELECT tq.question_number, tq.max_marks, stm.obtained_marks
        FROM test_questions tq
        LEFT JOIN student_test_marks stm
               ON tq.question_id = stm.question_id
              AND stm.moodle_id  = %s
              AND stm.test_id    = %s
              AND stm.attempt_no = %s
        WHERE tq.test_id = %s
        ORDER BY tq.question_number
    """, (moodle_id, test_id, attempt_no, test_id))
    marks_detail = cursor.fetchall()

    # Calculate totals in Python
    total_max      = sum(row['max_marks']            for row in marks_detail)
    total_obtained = sum(row['obtained_marks'] or 0  for row in marks_detail)
    percentage     = round(total_obtained / total_max * 100, 2) if total_max > 0 else 0

    cursor.close()
    conn.close()

    return render_template('view_marks_detail.html',
                           student=student, test=test, marks_detail=marks_detail,
                           total_max=total_max, total_obtained=total_obtained,
                           percentage=percentage, attempt_no=attempt_no)


# ============================================================
#  ATTENDANCE ROUTES  (admin only)
# ============================================================

@app.route('/admin/add_attendance', methods=['GET', 'POST'])
def add_attendance():
    """
    GET  – show the Add Attendance form with a student dropdown.
    POST – INSERT one daily attendance record (Present/Absent).
           The UNIQUE KEY (moodle_id, date) prevents duplicate entries.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        moodle_id = request.form.get('moodle_id')
        date      = request.form.get('date')    # string from <input type="date">
        status    = request.form.get('status')  # 'Present' or 'Absent'

        if not all([moodle_id, date, status]):
            flash('All fields are required', 'danger')
            cursor.execute("SELECT * FROM students ORDER BY moodle_id")
            students = cursor.fetchall()
            cursor.close()
            conn.close()
            return render_template('add_attendance.html', students=students)

        try:
            cursor.execute(
                "INSERT INTO attendance_daily (moodle_id, date, status) VALUES (%s, %s, %s)",
                (moodle_id, date, status)
            )
            conn.commit()
            flash('Attendance added successfully!', 'success')
            cursor.close()
            conn.close()
            return redirect(url_for('view_attendance'))

        except mysql.connector.IntegrityError:
            # UNIQUE KEY (moodle_id, date) was violated – record already exists for that day
            flash('Attendance for this student on this date already exists', 'danger')
            cursor.execute("SELECT * FROM students ORDER BY moodle_id")
            students = cursor.fetchall()
            cursor.close()
            conn.close()
            return render_template('add_attendance.html', students=students)

        except Exception as e:
            flash(f'Error adding attendance: {str(e)}', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('add_attendance'))

    # GET – load student list for the dropdown
    cursor.execute("SELECT * FROM students ORDER BY moodle_id")
    students = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('add_attendance.html', students=students)


@app.route('/admin/update_attendance_percentage', methods=['GET', 'POST'])
def update_attendance_percentage():
    """
    GET  – show all students with their current attendance %.
    POST – update (or insert) the attendance_percentage for one student.
           Uses INSERT … ON DUPLICATE KEY UPDATE (upsert pattern).
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        moodle_id          = request.form.get('moodle_id')
        attendance_percent = request.form.get('attendance_percent')

        if not moodle_id or attendance_percent is None:
            flash('All fields are required', 'danger')
            cursor.execute(
                "SELECT s.*, ap.attendance_percent FROM students s "
                "LEFT JOIN attendance_percentage ap ON s.moodle_id = ap.moodle_id "
                "ORDER BY s.moodle_id"
            )
            students = cursor.fetchall()
            cursor.close()
            conn.close()
            return render_template('update_attendance_percentage.html', students=students)

        try:
            attendance_percent = float(attendance_percent)
            if attendance_percent < 0 or attendance_percent > 100:
                raise ValueError()  # percentage must be 0–100

            # Upsert: insert if no row exists, update if it does
            cursor.execute("""
                INSERT INTO attendance_percentage (moodle_id, attendance_percent)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE attendance_percent = %s
            """, (moodle_id, attendance_percent, attendance_percent))
            conn.commit()
            flash('Attendance percentage updated successfully!', 'success')
            cursor.close()
            conn.close()
            return redirect(url_for('view_attendance'))

        except ValueError:
            flash('Attendance percentage must be between 0 and 100', 'danger')
            cursor.execute(
                "SELECT s.*, ap.attendance_percent FROM students s "
                "LEFT JOIN attendance_percentage ap ON s.moodle_id = ap.moodle_id "
                "ORDER BY s.moodle_id"
            )
            students = cursor.fetchall()
            cursor.close()
            conn.close()
            return render_template('update_attendance_percentage.html', students=students)

        except Exception as e:
            flash(f'Error updating attendance: {str(e)}', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('update_attendance_percentage'))

    # GET – load students with their current percentages
    cursor.execute(
        "SELECT s.*, ap.attendance_percent FROM students s "
        "LEFT JOIN attendance_percentage ap ON s.moodle_id = ap.moodle_id "
        "ORDER BY s.moodle_id"
    )
    students = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('update_attendance_percentage.html', students=students)


@app.route('/admin/view_attendance')
def view_attendance():
    """
    Show two tables:
      1. Daily attendance records (date, student, Present/Absent)
      2. Attendance percentage per student
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    # Daily records – most recent dates first
    cursor.execute("""
        SELECT ad.*, s.name, s.roll_no
        FROM attendance_daily ad
        JOIN students s ON ad.moodle_id = s.moodle_id
        ORDER BY ad.date DESC, ad.moodle_id
    """)
    daily_records = cursor.fetchall()

    # Percentage records – LEFT JOIN so students with no % entry still appear
    cursor.execute("""
        SELECT s.moodle_id, s.roll_no, s.name, ap.attendance_percent
        FROM students s
        LEFT JOIN attendance_percentage ap ON s.moodle_id = ap.moodle_id
        ORDER BY s.moodle_id
    """)
    percentage_records = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('view_attendance.html',
                           daily_records=daily_records,
                           percentage_records=percentage_records)


@app.route('/admin/delete_attendance/<int:record_id>')
def delete_attendance(record_id):
    """Delete a single daily attendance record by its record_id (primary key)."""
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM attendance_daily WHERE record_id = %s", (record_id,))
            conn.commit()
            flash('Attendance record deleted successfully!', 'success')
        except Exception as e:
            flash(f'Error deleting attendance: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    else:
        flash('Database connection error', 'danger')

    return redirect(url_for('view_attendance'))


# ============================================================
#  ANALYSIS DASHBOARD  (admin only)
# ============================================================

@app.route('/admin/analysis')
def analysis():
    """
    Fetch all data needed for the Chart.js graphs on analysis.html:
      - marks_data          – per student per test per attempt totals
      - attendance_summary  – total Present vs Absent counts
      - attendance_timeline – daily present/absent counts over time
      - attendance_percentages – per-student attendance %
      - test_stats          – avg/max/min per test
      - system stats        – 4 summary counts
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    # Marks grouped by student + test + attempt (for bar charts)
    cursor.execute("""
        SELECT s.name, s.moodle_id, t.test_name, stm.attempt_no,
               SUM(stm.obtained_marks) AS total_obtained,
               SUM(tq.max_marks)       AS total_max
        FROM student_test_marks stm
        JOIN students s        ON stm.moodle_id   = s.moodle_id
        JOIN tests t           ON stm.test_id     = t.test_id
        JOIN test_questions tq ON stm.question_id = tq.question_id
        GROUP BY s.moodle_id, t.test_id, stm.attempt_no
        ORDER BY s.name, t.test_name, stm.attempt_no
    """)
    marks_data = cursor.fetchall()

    # Overall attendance count (for pie chart: Present vs Absent)
    cursor.execute("""
        SELECT status, COUNT(*) AS count
        FROM attendance_daily
        GROUP BY status
    """)
    attendance_summary = cursor.fetchall()

    # Daily attendance counts over time (for line chart)
    cursor.execute("""
        SELECT date,
               SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS present_count,
               SUM(CASE WHEN status='Absent'  THEN 1 ELSE 0 END) AS absent_count
        FROM attendance_daily
        GROUP BY date
        ORDER BY date
    """)
    attendance_timeline = cursor.fetchall()

    # Per-student attendance percentage (for bar chart)
    cursor.execute("""
        SELECT s.name, ap.attendance_percent
        FROM attendance_percentage ap
        JOIN students s ON ap.moodle_id = s.moodle_id
        ORDER BY s.name
    """)
    attendance_percentages = cursor.fetchall()

    # Per-test statistics: attempt count, avg/max/min total marks
    cursor.execute("""
        SELECT t.test_name,
               COUNT(DISTINCT total_marks.moodle_id, total_marks.attempt_no) AS num_attempts,
               AVG(total_marks.total_obtained) AS avg_marks,
               MAX(total_marks.total_obtained) AS max_marks,
               MIN(total_marks.total_obtained) AS min_marks
        FROM tests t
        LEFT JOIN (
            SELECT test_id, moodle_id, attempt_no, SUM(obtained_marks) AS total_obtained
            FROM student_test_marks
            GROUP BY test_id, moodle_id, attempt_no
        ) AS total_marks ON t.test_id = total_marks.test_id
        GROUP BY t.test_id
    """)
    test_stats = cursor.fetchall()

    # System-wide summary counts
    cursor.execute("SELECT COUNT(*) AS count FROM students")
    total_students = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) AS count FROM tests")
    total_tests = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(DISTINCT moodle_id, test_id, attempt_no) AS count FROM student_test_marks")
    total_attempts = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) AS count FROM attendance_daily")
    total_attendance_entries = cursor.fetchone()['count']

    cursor.close()
    conn.close()

    return render_template('analysis.html',
                           marks_data=marks_data,
                           attendance_summary=attendance_summary,
                           attendance_timeline=attendance_timeline,
                           attendance_percentages=attendance_percentages,
                           test_stats=test_stats,
                           total_students=total_students,
                           total_tests=total_tests,
                           total_attempts=total_attempts,
                           total_attendance_entries=total_attendance_entries)


@app.route('/admin/test_analysis/<int:test_id>')
def test_analysis(test_id):
    """
    Detailed analysis for a single test:
      - marks_data – per-student per-attempt totals for this test
      - stats      – overall avg/max/min for this test
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tests WHERE test_id = %s", (test_id,))
    test = cursor.fetchone()

    if not test:
        flash('Test not found', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('view_tests'))

    # Marks for this specific test only
    cursor.execute("""
        SELECT s.name, s.moodle_id, stm.attempt_no,
               SUM(stm.obtained_marks) AS total_obtained,
               SUM(tq.max_marks)       AS total_max
        FROM student_test_marks stm
        JOIN students s        ON stm.moodle_id   = s.moodle_id
        JOIN test_questions tq ON stm.question_id = tq.question_id
        WHERE stm.test_id = %s
        GROUP BY s.moodle_id, stm.attempt_no
        ORDER BY s.name, stm.attempt_no
    """, (test_id,))
    marks_data = cursor.fetchall()

    # Aggregate stats for this test
    cursor.execute("""
        SELECT COUNT(DISTINCT moodle_id, attempt_no) AS total_attempts,
               AVG(total_marks.total_obtained)       AS avg_marks,
               MAX(total_marks.total_obtained)       AS max_marks,
               MIN(total_marks.total_obtained)       AS min_marks
        FROM (
            SELECT moodle_id, attempt_no, SUM(obtained_marks) AS total_obtained
            FROM student_test_marks
            WHERE test_id = %s
            GROUP BY moodle_id, attempt_no
        ) AS total_marks
    """, (test_id,))
    stats = cursor.fetchone()

    cursor.close()
    conn.close()

    return render_template('test_analysis.html', test=test, marks_data=marks_data, stats=stats)


# ============================================================
#  RAW DATA EXCEL EXPORT  (admin only)
# ============================================================

@app.route('/admin/export_excel', methods=['GET', 'POST'])
def export_excel():
    """
    GET  – show the export options form (checkboxes for each sheet).
    POST – build a multi-sheet openpyxl workbook from the checked
           options and stream it as a .xlsx download.

    Uses openpyxl (not XlsxWriter) because this is a simple
    data dump without charts.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        # Read which checkboxes were ticked (present in form = True)
        export_students          = 'students'          in request.form
        export_tests             = 'tests'             in request.form
        export_test_questions    = 'test_questions'    in request.form
        export_marks             = 'marks'             in request.form
        export_attendance_daily  = 'attendance_daily'  in request.form
        export_attendance_percent= 'attendance_percent'in request.form
        export_faculty           = 'faculty'           in request.form

        conn = get_db_connection()
        if not conn:
            flash('Database connection error', 'danger')
            return redirect(url_for('admin_dashboard'))

        cursor = conn.cursor(dictionary=True)

        # Create a new workbook and remove the default empty sheet
        wb = Workbook()
        wb.remove(wb.active)

        # Shared header cell style: blue background, white bold text, centred
        header_fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font      = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # ── Students sheet ────────────────────────────────────
        if export_students:
            ws = wb.create_sheet("Students")
            ws.append(['Moodle ID', 'Roll No', 'Name', 'Moodle Email', 'Phone'])
            for cell in ws[1]:  # style the header row
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_alignment
            cursor.execute("SELECT * FROM students ORDER BY moodle_id")
            for row in cursor.fetchall():
                ws.append([row['moodle_id'], row['roll_no'], row['name'],
                           row['moodle_email'], row['phone']])

        # ── Tests sheet ───────────────────────────────────────
        if export_tests:
            ws = wb.create_sheet("Tests")
            ws.append(['Test ID', 'Test Name', 'Created At'])
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_alignment
            cursor.execute("SELECT * FROM tests ORDER BY test_id")
            for row in cursor.fetchall():
                ws.append([row['test_id'], row['test_name'],
                           row['created_at'].strftime('%Y-%m-%d %H:%M:%S') if row['created_at'] else ''])

        # ── Test Questions sheet ──────────────────────────────
        if export_test_questions:
            ws = wb.create_sheet("Test Questions")
            ws.append(['Question ID', 'Test ID', 'Test Name', 'Question Number', 'Max Marks'])
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_alignment
            cursor.execute("""
                SELECT tq.*, t.test_name FROM test_questions tq
                JOIN tests t ON tq.test_id = t.test_id
                ORDER BY tq.test_id, tq.question_number
            """)
            for row in cursor.fetchall():
                ws.append([row['question_id'], row['test_id'], row['test_name'],
                           row['question_number'], row['max_marks']])

        # ── Student Test Marks sheet ──────────────────────────
        if export_marks:
            ws = wb.create_sheet("Student Test Marks")
            ws.append(['Moodle ID', 'Roll No', 'Name', 'Test Name',
                       'Attempt No', 'Question No', 'Max Marks', 'Obtained Marks'])
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_alignment
            cursor.execute("""
                SELECT s.moodle_id, s.roll_no, s.name, t.test_name, stm.attempt_no,
                       tq.question_number, tq.max_marks, stm.obtained_marks
                FROM student_test_marks stm
                JOIN students s        ON stm.moodle_id   = s.moodle_id
                JOIN tests t           ON stm.test_id     = t.test_id
                JOIN test_questions tq ON stm.question_id = tq.question_id
                ORDER BY s.moodle_id, t.test_id, stm.attempt_no, tq.question_number
            """)
            for row in cursor.fetchall():
                ws.append([row['moodle_id'], row['roll_no'], row['name'], row['test_name'],
                           row['attempt_no'], row['question_number'],
                           row['max_marks'], row['obtained_marks']])

        # ── Attendance Daily sheet ────────────────────────────
        if export_attendance_daily:
            ws = wb.create_sheet("Attendance Daily")
            ws.append(['Moodle ID', 'Roll No', 'Name', 'Date', 'Status'])
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_alignment
            cursor.execute("""
                SELECT ad.moodle_id, s.roll_no, s.name, ad.date, ad.status
                FROM attendance_daily ad
                JOIN students s ON ad.moodle_id = s.moodle_id
                ORDER BY ad.date DESC, ad.moodle_id
            """)
            for row in cursor.fetchall():
                ws.append([row['moodle_id'], row['roll_no'], row['name'],
                           row['date'].strftime('%Y-%m-%d') if row['date'] else '', row['status']])

        # ── Attendance Percentage sheet ───────────────────────
        if export_attendance_percent:
            ws = wb.create_sheet("Attendance Percentage")
            ws.append(['Moodle ID', 'Roll No', 'Name', 'Attendance Percent'])
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_alignment
            cursor.execute("""
                SELECT s.moodle_id, s.roll_no, s.name, ap.attendance_percent
                FROM students s
                LEFT JOIN attendance_percentage ap ON s.moodle_id = ap.moodle_id
                ORDER BY s.moodle_id
            """)
            for row in cursor.fetchall():
                ws.append([row['moodle_id'], row['roll_no'], row['name'],
                           row['attendance_percent'] or 0])

        # ── Faculty sheet ─────────────────────────────────────
        if export_faculty:
            ws = wb.create_sheet("Faculty")
            ws.append(['Faculty ID', 'Username', 'Role'])
            for cell in ws[1]:
                cell.fill = header_fill; cell.font = header_font; cell.alignment = header_alignment
            cursor.execute("SELECT faculty_id, username, role FROM faculty ORDER BY faculty_id")
            for row in cursor.fetchall():
                ws.append([row['faculty_id'], row['username'], row['role']])

        cursor.close()
        conn.close()

        # Write workbook into a BytesIO buffer (no temp file on disk)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)  # rewind so Flask reads from the beginning

        # Timestamp in filename so each download is uniquely named
        filename = f'student_assessment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    # GET – show the export options form
    return render_template('export_excel.html')


# ============================================================
#  FACULTY MANAGEMENT  (admin only)
# ============================================================

@app.route('/admin/add_faculty', methods=['GET', 'POST'])
def add_faculty():
    """
    GET  – show the Add Faculty form.
    POST – hash the password and INSERT a new faculty row with role='faculty'.
           Admins are created separately via create_admin.py.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not all([username, password]):
            flash('All fields are required', 'danger')
            return render_template('add_faculty.html')

        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            try:
                # Hash the password before storing (never store plain text)
                hashed_password = generate_password_hash(password)
                cursor.execute(
                    "INSERT INTO faculty (username, password, role) VALUES (%s, %s, 'faculty')",
                    (username, hashed_password)
                )
                conn.commit()
                flash('Faculty added successfully!', 'success')
                cursor.close()
                conn.close()
                return redirect(url_for('view_faculty'))
            except mysql.connector.IntegrityError:
                # username UNIQUE constraint violated
                flash('Username already exists', 'danger')
                cursor.close()
                conn.close()
            except Exception as e:
                flash(f'Error adding faculty: {str(e)}', 'danger')
                cursor.close()
                conn.close()
        else:
            flash('Database connection error', 'danger')

    return render_template('add_faculty.html')


@app.route('/admin/view_faculty')
def view_faculty():
    """List all faculty accounts (role='faculty', excludes admins)."""
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    cursor = conn.cursor(dictionary=True)
    # Only show faculty, not admins; newest first
    cursor.execute("SELECT * FROM faculty WHERE role = 'faculty' ORDER BY faculty_id DESC")
    faculty_list = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('view_faculty.html', faculty_list=faculty_list)


@app.route('/admin/delete_faculty/<int:faculty_id>')
def delete_faculty(faculty_id):
    """
    Delete a faculty account.
    The AND role='faculty' guard prevents accidentally deleting an admin.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            # Safety check: only delete if role is 'faculty' (not 'admin')
            cursor.execute(
                "DELETE FROM faculty WHERE faculty_id = %s AND role = 'faculty'",
                (faculty_id,)
            )
            conn.commit()
            flash('Faculty deleted successfully!', 'success')
        except Exception as e:
            flash(f'Error deleting faculty: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    else:
        flash('Database connection error', 'danger')

    return redirect(url_for('view_faculty'))


# ============================================================
#  ANALYTICS EXPORT WITH CHARTS  (admin only)
# ============================================================

@app.route('/admin/export_analytics')
def export_analytics():
    """
    Calls excel_manager.export_analytics() to build a rich 5-sheet
    workbook with embedded bar/pie charts using XlsxWriter,
    then streams it as a .xlsx download.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        output = excel_manager.export_analytics(conn)  # returns BytesIO
        filename = f'analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))
    finally:
        conn.close()  # always close the connection even if export fails


# ============================================================
#  EXCEL IMPORT  (admin only)
# ============================================================

@app.route('/admin/import_excel', methods=['GET', 'POST'])
def import_excel():
    """
    GET  – show the import upload form.
    POST – receive the uploaded .xlsx file, pass it to
           excel_manager.import_workbook(), and display results.

    The import function handles sheets named 'Students',
    'Attendance', and 'Marks'. Other sheets are ignored.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        file = request.files.get('excel_file')  # from <input type="file" name="excel_file">

        if not file or file.filename == '':
            flash('Please select an Excel file to upload.', 'danger')
            return render_template('import_excel.html')

        # Only accept Excel formats
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Only .xlsx / .xls files are supported.', 'danger')
            return render_template('import_excel.html')

        conn = get_db_connection()
        if not conn:
            flash('Database connection error', 'danger')
            return render_template('import_excel.html')

        try:
            # Pass the file stream directly – no need to save to disk
            results = excel_manager.import_workbook(file.stream, conn)
            if not results:
                flash('No recognised sheets found. Use sheets named: Students, Attendance, Marks.',
                      'warning')
            # Pass results to template to show inserted/skipped counts and errors
            return render_template('import_excel.html', results=results)
        except Exception as e:
            flash(f'Import failed: {str(e)}', 'danger')
            return render_template('import_excel.html')
        finally:
            conn.close()

    # GET – show the upload form
    return render_template('import_excel.html')


@app.route('/admin/import_template')
def import_template():
    """
    Generate and download a blank import template (.xlsx) with
    3 pre-formatted sheets: Students, Attendance, Marks.
    Users fill this in and upload it via import_excel.
    """
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    output = excel_manager.generate_import_template()  # returns BytesIO
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='import_template.xlsx'
    )


# ============================================================
#  ADVANCED ANALYTICS ROUTES  (admin only)
# ============================================================

@app.route('/admin/advanced-analytics')
def advanced_analytics():
    """Main analytics dashboard – overall, marks, attendance, filters."""
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        overall = analytics.get_overall_stats(conn)
        marks   = analytics.get_marks_analysis(conn)
        att     = analytics.get_attendance_analysis(conn)
    finally:
        conn.close()

    return render_template('advanced_analytics.html',
                           overall=overall, marks=marks, att=att)


@app.route('/admin/api/marks-analysis')
def api_marks_analysis():
    """JSON endpoint – marks data for Chart.js."""
    if not is_admin():
        from flask import jsonify
        return jsonify({'error': 'unauthorized'}), 403
    from flask import jsonify
    conn = get_db_connection()
    data = analytics.get_marks_analysis(conn)
    conn.close()
    return jsonify(data)


@app.route('/admin/api/attendance-analysis')
def api_attendance_analysis():
    """JSON endpoint – attendance data for Chart.js."""
    if not is_admin():
        from flask import jsonify
        return jsonify({'error': 'unauthorized'}), 403
    from flask import jsonify
    conn = get_db_connection()
    data = analytics.get_attendance_analysis(conn)
    conn.close()
    return jsonify(data)


@app.route('/admin/api/top-students')
def api_top_students():
    """JSON endpoint – filtered student list."""
    if not is_admin():
        from flask import jsonify
        return jsonify({'error': 'unauthorized'}), 403
    from flask import jsonify
    filter_type = request.args.get('filter', 'top10')
    conn = get_db_connection()
    data = analytics.get_filtered_students(conn, filter_type)
    conn.close()
    return jsonify(data)


@app.route('/admin/export-analysis')
def export_analysis():
    """Download the full 5-sheet analytics Excel file."""
    if not is_admin():
        flash('Access denied. Admin only.', 'danger')
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'danger')
        return redirect(url_for('admin_dashboard'))

    try:
        output = analytics_export.export_full_analysis(conn)
        filename = f'class_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'danger')
        return redirect(url_for('advanced_analytics'))
    finally:
        conn.close()


# ============================================================
#  APP ENTRY POINT
# ============================================================

if __name__ == '__main__':
    # debug=True enables auto-reload on code changes and shows
    # detailed error pages. Turn this OFF in production.
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '').lower() in {'1', 'true', 'yes'}
    app.run(host='0.0.0.0', port=port, debug=debug)
