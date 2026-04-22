"""
============================================================
excel_manager.py  –  Excel Import / Export Manager
============================================================
Two responsibilities:
  EXPORT  – build a rich analytics workbook with embedded
            charts using XlsxWriter (write-only, fast).
  IMPORT  – parse an uploaded workbook and bulk-upsert
            students, attendance, and marks using openpyxl
            (read-capable).

Why two libraries?
  • XlsxWriter  → supports embedded charts (openpyxl does not)
  • openpyxl    → can READ existing .xlsx files (XlsxWriter cannot)
============================================================
"""

import io                          # BytesIO – in-memory file buffer (no disk write needed)
import xlsxwriter                  # Used for EXPORT – supports charts
from openpyxl import load_workbook # Used for IMPORT – can read .xlsx files
from datetime import datetime, date # Used to handle date parsing in import


# ============================================================
#  EXPORT  –  Analytics workbook with charts
# ============================================================

def _col_width(ws, col, width):
    """Helper: set a single column width in an XlsxWriter worksheet."""
    ws.set_column(col, col, width)


def export_analytics(conn):
    """
    Build a 5-sheet analytics Excel workbook and return it as BytesIO.

    Sheets produced:
      1. Summary        – stat cards + horizontal bar chart (avg score per student)
      2. Marks Analysis – colour-coded marks table + column chart per test
      3. Attendance     – per-student table + pie chart (Present vs Absent)
      4. Students       – plain student master list
      5. Raw Marks      – every question mark for every attempt

    Parameters:
      conn – active MySQL connection (caller is responsible for closing it)

    Returns:
      io.BytesIO positioned at 0, ready to pass to Flask's send_file()
    """

    # ── Create in-memory workbook ─────────────────────────────
    output = io.BytesIO()  # buffer that acts like a file but lives in RAM
    wb = xlsxwriter.Workbook(output, {'in_memory': True})  # write into the buffer

    # ── Define reusable cell formats ──────────────────────────
    # hdr       – blue background, white bold text, centred (used for column headers)
    hdr = wb.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF',
                         'align': 'center', 'valign': 'vcenter', 'border': 1})

    # title_fmt – large dark-blue text for sheet titles
    title_fmt = wb.add_format({'bold': True, 'font_size': 16, 'font_color': '#1F3864'})

    # sub_fmt   – medium blue text for section headings
    sub_fmt   = wb.add_format({'bold': True, 'font_size': 12, 'font_color': '#2F5496'})

    # stat_lbl  – light blue background label cell (stat card label row)
    stat_lbl  = wb.add_format({'bold': True, 'font_size': 11, 'bg_color': '#D9E1F2', 'border': 1})

    # stat_val  – blue background value cell (big number in stat card)
    stat_val  = wb.add_format({'font_size': 14, 'bold': True, 'align': 'center',
                               'bg_color': '#4472C4', 'font_color': '#FFFFFF', 'border': 1})

    # cell_fmt  – standard data cell (centred, bordered)
    cell_fmt  = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})

    # pct_fmt   – percentage cell showing two decimal places e.g. "87.50%"
    pct_fmt   = wb.add_format({'num_format': '0.00"%"', 'align': 'center', 'border': 1})

    # date_fmt  – date cell formatted as YYYY-MM-DD (not used directly but kept for reference)
    date_fmt  = wb.add_format({'num_format': 'yyyy-mm-dd', 'align': 'center', 'border': 1})

    # green_fmt – light green cell (score >= 50% or attendance >= 75%)
    green_fmt = wb.add_format({'bg_color': '#E2EFDA', 'align': 'center', 'border': 1})

    # red_fmt   – light red/orange cell (score < 50% or attendance < 75%)
    red_fmt   = wb.add_format({'bg_color': '#FCE4D6', 'align': 'center', 'border': 1})

    # Open a DB cursor that returns rows as dicts (column name → value)
    cur = conn.cursor(dictionary=True)

    # ──────────────────────────────────────────────────────────
    #  SHEET 1: Summary
    #  Shows 4 stat cards and a bar chart of avg score per student
    # ──────────────────────────────────────────────────────────
    ws = wb.add_worksheet('Summary')
    ws.set_zoom(90)  # zoom out slightly so more fits on screen

    # Report title and generation timestamp
    ws.write('B2', 'Student Performance Analysis and Visualization Dashboard – Analytics Report', title_fmt)
    ws.write('B3', f'Generated: {datetime.now().strftime("%d %b %Y  %H:%M")}', sub_fmt)

    # ── Fetch counts for the 4 stat cards ────────────────────
    cur.execute("SELECT COUNT(*) AS c FROM students")
    n_stu = cur.fetchone()['c']   # total number of students

    cur.execute("SELECT COUNT(*) AS c FROM tests")
    n_tst = cur.fetchone()['c']   # total number of tests

    # COUNT DISTINCT on 3 columns = unique student+test+attempt combinations
    cur.execute("SELECT COUNT(DISTINCT moodle_id,test_id,attempt_no) AS c FROM student_test_marks")
    n_att = cur.fetchone()['c']   # total test attempts recorded

    cur.execute("SELECT COUNT(*) AS c FROM attendance_daily")
    n_atn = cur.fetchone()['c']   # total daily attendance entries

    # ── Write stat cards (label row 5, value row 6) ──────────
    stats = [
        ('Total Students',     n_stu),
        ('Total Tests',        n_tst),
        ('Test Attempts',      n_att),
        ('Attendance Records', n_atn),
    ]
    for i, (lbl, val) in enumerate(stats):
        c = 1 + i * 2          # place each card 2 columns apart (B, D, F, H)
        ws.write(5, c, lbl, stat_lbl)   # row 6 in Excel (0-indexed row 5)
        ws.write(6, c, val, stat_val)   # row 7 in Excel (0-indexed row 6)
        ws.set_column(c, c, 18)         # widen the column so text fits

    # ── Bar chart: average score % per student (top 20) ──────
    ws.write(9, 1, 'Average Score per Student (Top 20)', sub_fmt)
    ws.write(10, 1, 'Moodle ID', hdr)   # column header for chart data table
    ws.write(10, 2, 'Avg %', hdr)

    # SQL: sum obtained / sum max * 100 grouped by student, top 20 by score
    cur.execute("""
        SELECT s.moodle_id, s.name,
               ROUND(SUM(stm.obtained_marks)/SUM(tq.max_marks)*100,2) AS avg_pct
        FROM student_test_marks stm
        JOIN students s ON stm.moodle_id=s.moodle_id
        JOIN test_questions tq ON stm.question_id=tq.question_id
        GROUP BY s.moodle_id ORDER BY avg_pct DESC LIMIT 20
    """)
    rows = cur.fetchall()

    # Write the data table that the chart will reference
    for r, row in enumerate(rows):
        ws.write(11+r, 1, row['moodle_id'], cell_fmt)  # student ID in col B
        ws.write(11+r, 2, row['avg_pct'],   pct_fmt)   # avg % in col C

    # Only create the chart if there is data to plot
    if rows:
        chart = wb.add_chart({'type': 'bar'})  # horizontal bar chart
        chart.add_series({
            'name': 'Avg Score %',
            # categories = the Y-axis labels (student IDs)
            'categories': ['Summary', 11, 1, 11+len(rows)-1, 1],
            # values = the bar lengths (avg percentages)
            'values':     ['Summary', 11, 2, 11+len(rows)-1, 2],
            'fill': {'color': '#4472C4'},  # blue bars
            'gap': 50,                     # gap between bars
        })
        chart.set_title({'name': 'Average Score % per Student'})
        chart.set_x_axis({'name': 'Score %'})
        chart.set_y_axis({'name': 'Student'})
        chart.set_size({'width': 600, 'height': 400})
        chart.set_style(10)               # built-in Excel chart style
        ws.insert_chart('E10', chart)     # place chart starting at cell E10

    # Set column widths for the data table
    ws.set_column(1, 1, 18)  # Moodle ID column
    ws.set_column(2, 2, 12)  # Avg % column

    # ──────────────────────────────────────────────────────────
    #  SHEET 2: Marks Analysis
    #  Per-student per-test totals, colour-coded, + column chart
    # ──────────────────────────────────────────────────────────
    ws2 = wb.add_worksheet('Marks Analysis')
    ws2.write(0, 0, 'Test-wise Marks Analysis', title_fmt)  # sheet title

    # Column headers for the marks table
    hdrs2 = ['Moodle ID', 'Name', 'Test Name', 'Attempt', 'Total Max', 'Total Obtained', 'Score %']
    for c, h in enumerate(hdrs2):
        ws2.write(2, c, h, hdr)                          # write header in row 3
        ws2.set_column(c, c, [14,22,24,10,12,16,10][c])  # set each column's width

    # SQL: aggregate marks per student per test per attempt
    cur.execute("""
        SELECT s.moodle_id, s.name, t.test_name, stm.attempt_no,
               SUM(tq.max_marks)        AS total_max,
               SUM(stm.obtained_marks)  AS total_obt
        FROM student_test_marks stm
        JOIN students s       ON stm.moodle_id  = s.moodle_id
        JOIN tests t          ON stm.test_id    = t.test_id
        JOIN test_questions tq ON stm.question_id = tq.question_id
        GROUP BY s.moodle_id, t.test_id, stm.attempt_no
        ORDER BY t.test_name, s.moodle_id, stm.attempt_no
    """)
    marks_rows = cur.fetchall()

    for r, row in enumerate(marks_rows):
        # Calculate score percentage; avoid division by zero
        pct = round(row['total_obt'] / row['total_max'] * 100, 2) if row['total_max'] else 0
        # Green if passing (>=50%), red if failing
        fmt = green_fmt if pct >= 50 else red_fmt

        ws2.write(3+r, 0, row['moodle_id'],        cell_fmt)
        ws2.write(3+r, 1, row['name'],             cell_fmt)
        ws2.write(3+r, 2, row['test_name'],        cell_fmt)
        ws2.write(3+r, 3, row['attempt_no'],       cell_fmt)
        ws2.write(3+r, 4, int(row['total_max']),   cell_fmt)
        ws2.write(3+r, 5, int(row['total_obt']),   cell_fmt)
        ws2.write(3+r, 6, pct,                     fmt)      # colour-coded score %

    # ── Column chart: average score per test ─────────────────
    # SQL: average of per-attempt percentages, grouped by test
    cur.execute("""
        SELECT t.test_name,
               ROUND(AVG(sub.pct), 2) AS avg_pct
        FROM (
            SELECT stm.test_id,
                   SUM(stm.obtained_marks) / SUM(tq.max_marks) * 100 AS pct
            FROM student_test_marks stm
            JOIN test_questions tq ON stm.question_id = tq.question_id
            GROUP BY stm.moodle_id, stm.test_id, stm.attempt_no
        ) sub
        JOIN tests t ON sub.test_id = t.test_id
        GROUP BY t.test_id ORDER BY t.test_id
    """)
    test_avgs = cur.fetchall()

    # Place the chart data table below the main marks table (with a 3-row gap)
    chart_start = 3 + len(marks_rows) + 3
    ws2.write(chart_start,   0, 'Test Name',   hdr)
    ws2.write(chart_start,   1, 'Avg Score %', hdr)
    for r, row in enumerate(test_avgs):
        ws2.write(chart_start+1+r, 0, row['test_name'], cell_fmt)
        ws2.write(chart_start+1+r, 1, row['avg_pct'],   pct_fmt)

    if test_avgs:
        chart2 = wb.add_chart({'type': 'column'})  # vertical column chart
        chart2.add_series({
            'name': 'Avg Score %',
            'categories': ['Marks Analysis', chart_start+1, 0, chart_start+len(test_avgs), 0],
            'values':     ['Marks Analysis', chart_start+1, 1, chart_start+len(test_avgs), 1],
            'fill': {'color': '#70AD47'},  # green bars
            'gap': 60,
        })
        chart2.set_title({'name': 'Average Score % per Test'})
        chart2.set_x_axis({'name': 'Test'})
        chart2.set_y_axis({'name': 'Avg Score %', 'min': 0, 'max': 100})
        chart2.set_size({'width': 500, 'height': 320})
        chart2.set_style(10)
        ws2.insert_chart('I3', chart2)  # place chart at column I, row 3

    # ──────────────────────────────────────────────────────────
    #  SHEET 3: Attendance
    #  Per-student present/absent counts + overall pie chart
    # ──────────────────────────────────────────────────────────
    ws3 = wb.add_worksheet('Attendance')
    ws3.write(0, 0, 'Attendance Analysis', title_fmt)

    hdrs3 = ['Moodle ID', 'Name', 'Present', 'Absent', 'Total', 'Attendance %']
    for c, h in enumerate(hdrs3):
        ws3.write(2, c, h, hdr)
        ws3.set_column(c, c, [14,22,10,10,10,14][c])

    # SQL: count Present and Absent days per student using conditional SUM
    cur.execute("""
        SELECT s.moodle_id, s.name,
               SUM(CASE WHEN ad.status='Present' THEN 1 ELSE 0 END) AS present_cnt,
               SUM(CASE WHEN ad.status='Absent'  THEN 1 ELSE 0 END) AS absent_cnt,
               COUNT(*) AS total_cnt
        FROM students s
        LEFT JOIN attendance_daily ad ON s.moodle_id = ad.moodle_id
        GROUP BY s.moodle_id ORDER BY s.moodle_id
    """)
    atn_rows = cur.fetchall()

    # Accumulators for the pie chart totals
    total_present = total_absent = 0

    for r, row in enumerate(atn_rows):
        # Attendance % = present days / total days * 100
        pct = round(row['present_cnt'] / row['total_cnt'] * 100, 2) if row['total_cnt'] else 0
        # Green if >= 75% (common passing threshold), red otherwise
        fmt = green_fmt if pct >= 75 else red_fmt

        ws3.write(3+r, 0, row['moodle_id'],              cell_fmt)
        ws3.write(3+r, 1, row['name'],                   cell_fmt)
        ws3.write(3+r, 2, int(row['present_cnt'] or 0),  cell_fmt)
        ws3.write(3+r, 3, int(row['absent_cnt']  or 0),  cell_fmt)
        ws3.write(3+r, 4, int(row['total_cnt']   or 0),  cell_fmt)
        ws3.write(3+r, 5, pct,                           fmt)

        # Add to running totals for the pie chart
        total_present += int(row['present_cnt'] or 0)
        total_absent  += int(row['absent_cnt']  or 0)

    # ── Pie chart data table (placed below the main table) ───
    pie_row = 3 + len(atn_rows) + 2   # leave a 2-row gap
    ws3.write(pie_row,   0, 'Status',  hdr);    ws3.write(pie_row,   1, 'Count', hdr)
    ws3.write(pie_row+1, 0, 'Present', cell_fmt); ws3.write(pie_row+1, 1, total_present, cell_fmt)
    ws3.write(pie_row+2, 0, 'Absent',  cell_fmt); ws3.write(pie_row+2, 1, total_absent,  cell_fmt)

    # Pie chart – green slice = Present, red slice = Absent
    pie = wb.add_chart({'type': 'pie'})
    pie.add_series({
        'name': 'Attendance',
        'categories': ['Attendance', pie_row+1, 0, pie_row+2, 0],  # labels: Present / Absent
        'values':     ['Attendance', pie_row+1, 1, pie_row+2, 1],  # counts
        'points': [
            {'fill': {'color': '#70AD47'}},  # green for Present
            {'fill': {'color': '#FF0000'}},  # red for Absent
        ],
    })
    pie.set_title({'name': 'Overall Attendance Distribution'})
    pie.set_style(10)
    pie.set_size({'width': 420, 'height': 300})
    ws3.insert_chart('H3', pie)  # place pie chart at column H, row 3

    # ──────────────────────────────────────────────────────────
    #  SHEET 4: Students
    #  Plain master list of all students
    # ──────────────────────────────────────────────────────────
    ws4 = wb.add_worksheet('Students')
    ws4.write(0, 0, 'Student Master List', title_fmt)

    hdrs4 = ['Moodle ID', 'Roll No', 'Name', 'Email', 'Phone']
    for c, h in enumerate(hdrs4):
        ws4.write(1, c, h, hdr)
        ws4.set_column(c, c, [14,14,24,30,14][c])

    cur.execute("SELECT * FROM students ORDER BY moodle_id")
    for r, row in enumerate(cur.fetchall()):
        ws4.write(2+r, 0, row['moodle_id'],    cell_fmt)
        ws4.write(2+r, 1, row['roll_no'],      cell_fmt)
        ws4.write(2+r, 2, row['name'],         cell_fmt)
        ws4.write(2+r, 3, row['moodle_email'], cell_fmt)
        ws4.write(2+r, 4, row['phone'],        cell_fmt)

    # ──────────────────────────────────────────────────────────
    #  SHEET 5: Raw Marks
    #  Every question mark for every student, test, and attempt
    # ──────────────────────────────────────────────────────────
    ws5 = wb.add_worksheet('Raw Marks')
    ws5.write(0, 0, 'Question-wise Marks (All Attempts)', title_fmt)

    hdrs5 = ['Moodle ID', 'Roll No', 'Name', 'Test', 'Attempt', 'Q No', 'Max', 'Obtained']
    for c, h in enumerate(hdrs5):
        ws5.write(1, c, h, hdr)
        ws5.set_column(c, c, [14,12,22,22,10,8,8,10][c])

    # SQL: join all 4 tables to get one row per question per attempt
    cur.execute("""
        SELECT s.moodle_id, s.roll_no, s.name, t.test_name,
               stm.attempt_no, tq.question_number, tq.max_marks, stm.obtained_marks
        FROM student_test_marks stm
        JOIN students s        ON stm.moodle_id   = s.moodle_id
        JOIN tests t           ON stm.test_id     = t.test_id
        JOIN test_questions tq ON stm.question_id = tq.question_id
        ORDER BY s.moodle_id, t.test_id, stm.attempt_no, tq.question_number
    """)
    for r, row in enumerate(cur.fetchall()):
        ws5.write(2+r, 0, row['moodle_id'],       cell_fmt)
        ws5.write(2+r, 1, row['roll_no'],         cell_fmt)
        ws5.write(2+r, 2, row['name'],            cell_fmt)
        ws5.write(2+r, 3, row['test_name'],       cell_fmt)
        ws5.write(2+r, 4, row['attempt_no'],      cell_fmt)
        ws5.write(2+r, 5, row['question_number'], cell_fmt)
        ws5.write(2+r, 6, row['max_marks'],       cell_fmt)
        ws5.write(2+r, 7, row['obtained_marks'],  cell_fmt)

    # ── Finalise and return ───────────────────────────────────
    cur.close()
    wb.close()       # flushes all data into the BytesIO buffer
    output.seek(0)   # rewind to the start so Flask can read it
    return output


# ============================================================
#  IMPORT  –  Parse uploaded Excel workbook
# ============================================================

def _cell_val(cell):
    """
    Helper: return a clean value from an openpyxl cell.
    Strings are stripped of whitespace; numbers are returned as-is.
    Empty cells return None.
    """
    v = cell.value
    if v is None:
        return None
    return str(v).strip() if not isinstance(v, (int, float)) else v


def import_students(ws, conn):
    """
    Import students from a worksheet into the `students` table.

    Expected column order (row 1 = header, data starts row 2):
      Col A: Moodle ID  |  Col B: Roll No  |  Col C: Name
      Col D: Email      |  Col E: Phone

    Uses INSERT … ON DUPLICATE KEY UPDATE so existing students
    are updated rather than causing an error.
    Also inserts a row in attendance_percentage (default 0%) if
    the student doesn't have one yet.

    Returns: (inserted_count, skipped_count, error_messages[])
    """
    cur = conn.cursor()
    inserted = skipped = 0
    errors = []

    # iter_rows(min_row=2) skips the header row
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # skip completely empty rows

        try:
            # Unpack first 5 columns; convert None → empty string for validation
            moodle_id, roll_no, name, email, phone = [
                str(v).strip() if v is not None else '' for v in row[:5]
            ]

            # All 5 fields are mandatory
            if not all([moodle_id, roll_no, name, email, phone]):
                errors.append(f'Row {i}: missing required field – skipped')
                skipped += 1
                continue

            # Upsert: insert new or update existing student
            cur.execute(
                "INSERT INTO students (moodle_id,roll_no,name,moodle_email,phone) "
                "VALUES (%s,%s,%s,%s,%s) "
                "ON DUPLICATE KEY UPDATE roll_no=%s, name=%s, moodle_email=%s, phone=%s",
                (moodle_id, roll_no, name, email, phone,
                 roll_no, name, email, phone)
            )

            # Ensure the student has an attendance_percentage row (INSERT IGNORE = skip if exists)
            cur.execute(
                "INSERT IGNORE INTO attendance_percentage (moodle_id,attendance_percent) VALUES (%s,0)",
                (moodle_id,)
            )
            inserted += 1

        except Exception as e:
            errors.append(f'Row {i}: {e}')
            skipped += 1

    conn.commit()   # save all changes in one transaction
    cur.close()
    return inserted, skipped, errors


def import_attendance(ws, conn):
    """
    Import daily attendance records from a worksheet.

    Expected column order:
      Col A: Moodle ID  |  Col B: Date (YYYY-MM-DD)  |  Col C: Status (Present/Absent)

    Accepts date values as Python date/datetime objects (Excel native)
    or as plain strings in YYYY-MM-DD format.
    Uses ON DUPLICATE KEY UPDATE so re-uploading the same date
    just overwrites the status instead of erroring.

    Returns: (inserted_count, skipped_count, error_messages[])
    """
    cur = conn.cursor()
    inserted = skipped = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # skip blank rows

        try:
            moodle_id = str(row[0]).strip() if row[0] is not None else ''
            raw_date  = row[1]   # could be a date object or a string
            # capitalize() turns 'present' → 'Present', 'ABSENT' → 'Absent'
            status    = str(row[2]).strip().capitalize() if row[2] is not None else ''

            if not all([moodle_id, raw_date, status]):
                errors.append(f'Row {i}: missing field – skipped')
                skipped += 1
                continue

            # Validate status value
            if status not in ('Present', 'Absent'):
                errors.append(f'Row {i}: status must be Present/Absent – got "{status}"')
                skipped += 1
                continue

            # Normalise date: Excel may give a datetime/date object or a string
            if isinstance(raw_date, (datetime, date)):
                # datetime → extract just the date part; date → use as-is
                date_val = raw_date.date() if isinstance(raw_date, datetime) else raw_date
            else:
                # String → parse it
                date_val = datetime.strptime(str(raw_date).strip(), '%Y-%m-%d').date()

            # Upsert: insert new record or update status if date already exists
            cur.execute(
                "INSERT INTO attendance_daily (moodle_id,date,status) VALUES (%s,%s,%s) "
                "ON DUPLICATE KEY UPDATE status=%s",
                (moodle_id, date_val, status, status)
            )
            inserted += 1

        except Exception as e:
            errors.append(f'Row {i}: {e}')
            skipped += 1

    conn.commit()
    cur.close()
    return inserted, skipped, errors


def import_marks(ws, conn):
    """
    Import student marks from a worksheet.

    Expected column order:
      Col A: Moodle ID  |  Col B: Test Name  |  Col C: Attempt No
      Col D: Q No       |  Col E: Obtained Marks

    IMPORTANT: The test and its questions must already exist in the
    database (created via the admin UI) before importing marks.

    Validation:
      • Test name must match an existing test exactly.
      • Question number must exist for that test.
      • Obtained marks are clamped to max_marks if exceeded.

    Returns: (inserted_count, skipped_count, error_messages[])
    """
    cur = conn.cursor(dictionary=True)
    inserted = skipped = 0
    errors = []

    # Pre-load all test names → IDs into a dict to avoid repeated DB lookups
    cur.execute("SELECT test_id, test_name FROM tests")
    test_map = {r['test_name']: r['test_id'] for r in cur.fetchall()}
    # e.g. {'Mid Term Test': 1, 'Final Exam': 2}

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # skip blank rows

        try:
            moodle_id  = str(row[0]).strip() if row[0] is not None else ''
            test_name  = str(row[1]).strip() if row[1] is not None else ''
            attempt_no = int(row[2]) if row[2] is not None else 1   # default attempt 1
            q_no       = int(row[3]) if row[3] is not None else None
            obtained   = int(row[4]) if row[4] is not None else 0

            # All key fields must be present
            if not all([moodle_id, test_name, q_no is not None]):
                errors.append(f'Row {i}: missing field – skipped')
                skipped += 1
                continue

            # Look up test_id from the pre-loaded map
            test_id = test_map.get(test_name)
            if not test_id:
                errors.append(f'Row {i}: test "{test_name}" not found – skipped')
                skipped += 1
                continue

            # Look up the question to get its question_id and max_marks
            cur.execute(
                "SELECT question_id, max_marks FROM test_questions "
                "WHERE test_id=%s AND question_number=%s",
                (test_id, q_no)
            )
            q = cur.fetchone()
            if not q:
                errors.append(f'Row {i}: Q{q_no} not found in test "{test_name}" – skipped')
                skipped += 1
                continue

            # Clamp obtained marks to max_marks (warn but don't skip)
            if obtained > q['max_marks']:
                errors.append(
                    f'Row {i}: obtained {obtained} > max {q["max_marks"]} for Q{q_no} – clamped'
                )
                obtained = q['max_marks']

            # Upsert: insert new mark or update if same student+test+attempt+question exists
            cur.execute(
                "INSERT INTO student_test_marks "
                "(moodle_id,test_id,attempt_no,question_id,obtained_marks) "
                "VALUES (%s,%s,%s,%s,%s) "
                "ON DUPLICATE KEY UPDATE obtained_marks=%s",
                (moodle_id, test_id, attempt_no, q['question_id'], obtained, obtained)
            )
            inserted += 1

        except Exception as e:
            errors.append(f'Row {i}: {e}')
            skipped += 1

    conn.commit()
    cur.close()
    return inserted, skipped, errors


def import_workbook(file_stream, conn):
    """
    Entry point for importing an uploaded Excel file.

    Reads the workbook from file_stream (Flask's request.files stream),
    then dispatches each sheet to the correct import function based on
    the sheet name (case-insensitive):
      'Students'   → import_students()
      'Attendance' → import_attendance()
      'Marks'      → import_marks()

    Sheets with other names are silently ignored.

    Returns:
      dict  { sheet_name: {'inserted': int, 'skipped': int, 'errors': list} }
    """
    # data_only=True reads cell values, not formulas
    wb = load_workbook(file_stream, data_only=True)
    results = {}

    # Map lowercase sheet names to their handler functions
    sheet_handlers = {
        'students':   import_students,
        'attendance': import_attendance,
        'marks':      import_marks,
    }

    for sheet_name in wb.sheetnames:
        key = sheet_name.strip().lower()   # normalise: 'Students' → 'students'
        if key in sheet_handlers:
            ins, skp, errs = sheet_handlers[key](wb[sheet_name], conn)
            results[sheet_name] = {
                'inserted': ins,
                'skipped':  skp,
                'errors':   errs,
            }
        # Sheets not in the map (e.g. 'Tests', 'Raw Marks') are skipped

    return results


def generate_import_template():
    """
    Build and return a blank import template as BytesIO.

    The template has 3 sheets with styled headers and one example
    data row each, so users know exactly what format to follow.

    Sheets:
      Students   – Moodle ID, Roll No, Name, Email, Phone
      Attendance – Moodle ID, Date, Status
      Marks      – Moodle ID, Test Name, Attempt No, Q No, Obtained Marks

    Returns: io.BytesIO positioned at 0
    """
    import xlsxwriter as xw   # local import (already imported at top, but explicit here)
    output = io.BytesIO()
    wb = xw.Workbook(output, {'in_memory': True})

    # Blue header format (same style as the export sheets)
    hdr = wb.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF',
                         'align': 'center', 'border': 1})
    # Grey italic format for the example data rows
    ex  = wb.add_format({'italic': True, 'font_color': '#595959'})

    # ── Students sheet ────────────────────────────────────────
    ws = wb.add_worksheet('Students')
    for c, h in enumerate(['Moodle ID', 'Roll No', 'Name', 'Email', 'Phone']):
        ws.write(0, c, h, hdr)
        ws.set_column(c, c, [14,12,24,30,14][c])  # set column widths
    # Example row so users understand the expected format
    ws.write(1, 0, 'M001', ex); ws.write(1, 1, 'R001', ex)
    ws.write(1, 2, 'John Doe', ex); ws.write(1, 3, 'john@example.com', ex)
    ws.write(1, 4, '9999999999', ex)

    # ── Attendance sheet ──────────────────────────────────────
    ws2 = wb.add_worksheet('Attendance')
    for c, h in enumerate(['Moodle ID', 'Date (YYYY-MM-DD)', 'Status (Present/Absent)']):
        ws2.write(0, c, h, hdr)
        ws2.set_column(c, c, [14,22,26][c])
    ws2.write(1, 0, 'M001', ex); ws2.write(1, 1, '2025-01-15', ex)
    ws2.write(1, 2, 'Present', ex)

    # ── Marks sheet ───────────────────────────────────────────
    ws3 = wb.add_worksheet('Marks')
    for c, h in enumerate(['Moodle ID', 'Test Name', 'Attempt No', 'Q No', 'Obtained Marks']):
        ws3.write(0, c, h, hdr)
        ws3.set_column(c, c, [14,24,12,8,16][c])
    ws3.write(1, 0, 'M001', ex); ws3.write(1, 1, 'Mid Term Test', ex)
    ws3.write(1, 2, 1, ex); ws3.write(1, 3, 1, ex); ws3.write(1, 4, 8, ex)

    wb.close()
    output.seek(0)  # rewind so Flask can read from the start
    return output
